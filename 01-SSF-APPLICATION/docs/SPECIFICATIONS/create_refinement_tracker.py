#!/usr/bin/env python3
"""
Create SSF Refinement Tracker & Session Planner Excel file
Compatible with Google Sheets import
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def style_header(cell, bg_color='366092', font_color='FFFFFF'):
    """Apply header styling"""
    cell.font = Font(bold=True, color=font_color, size=11)
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def style_section_header(cell, bg_color='D9E1F2'):
    """Apply section header styling"""
    cell.font = Font(bold=True, size=10)
    cell.fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')

def create_workbook():
    """Create the complete workbook with all sheets"""
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create all sheets
    create_dashboard_sheet(wb)
    create_session_log_sheet(wb)
    create_burndown_sheet(wb)
    create_burnup_sheet(wb)
    create_epic_progress_sheet(wb)
    create_session_calendar_sheet(wb)
    create_quick_reference_sheet(wb)
    
    return wb

def create_dashboard_sheet(wb):
    """Create DASHBOARD sheet"""
    ws = wb.create_sheet('DASHBOARD')
    
    ws['A1'] = 'REFINEMENT TRACKER - DASHBOARD'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws.merge_cells('A1:G1')
    
    ws['A2'] = 'Last Updated'
    ws['B2'] = '=TODAY()'
    ws['C2'] = 'Current Sprint'
    ws['D2'] = 1
    
    ws['A4'] = 'QUICK METRICS'
    style_section_header(ws['A4'])
    ws.merge_cells('A4:B4')
    
    ws['A6'] = 'Total Stories'
    ws['C6'] = 91
    ws['E6'] = 'Status'
    ws['F6'] = '=IF(F9="","-",IF(F9>=6.5,"🟢 On Track",IF(F9>=4,"🟡 At Risk","🔴 Critical")))'
    
    ws['A7'] = 'To Refine (Open)'
    ws['C7'] = 84
    ws['E7'] = 'Risk Level'
    ws['F7'] = '=IF(C10>60,"🔴 HIGH",IF(C10>40,"🟡 MEDIUM","🟢 LOW"))'
    
    ws['A8'] = 'Already Refined (Ready)'
    ws['C8'] = 7
    
    ws['A9'] = 'Completed So Far'
    ws['C9'] = '=SUM(\'SESSION LOG\'!F:F)'
    
    ws['A10'] = 'Remaining'
    ws['C10'] = '=IF(ISBLANK(C9),84,84-C9)'
    
    ws['A11'] = 'Progress %'
    ws['C11'] = '=IF(C9=0,0,C9/84*100)'
    
    ws['E9'] = 'Current Velocity'
    ws['F9'] = '=IF(COUNTA(\'SESSION LOG\'!F2:F16)=0,0,AVERAGE(\'SESSION LOG\'!F2:F16))'
    
    ws['E10'] = 'Target Velocity'
    ws['F10'] = 6.5
    
    ws['E11'] = 'Projected Completion'
    ws['F11'] = '=IF(OR(F9=0,F9=""),"TBD","Sprint "&ROUNDUP(C10/F9/3+1,0))'
    
    ws['A13'] = 'SESSION SUMMARY (LAST 3)'
    style_section_header(ws['A13'])
    ws.merge_cells('A13:G13')
    
    headers = ['Session', 'Sprint', 'Date', 'Stories Refined', 'Cumulative', 'Status']
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=14, column=col)
        cell.value = header
        style_header(cell)
    
    ws['B15'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=1,"",INDEX(\'SESSION LOG\'!A:A,COUNTA(\'SESSION LOG\'!A:A)))'
    ws['C15'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=1,"",INDEX(\'SESSION LOG\'!B:B,COUNTA(\'SESSION LOG\'!A:A)))'
    ws['D15'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=1,"",INDEX(\'SESSION LOG\'!C:C,COUNTA(\'SESSION LOG\'!A:A)))'
    ws['E15'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=1,"",INDEX(\'SESSION LOG\'!F:F,COUNTA(\'SESSION LOG\'!A:A)))'
    ws['F15'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=1,"",INDEX(\'SESSION LOG\'!G:G,COUNTA(\'SESSION LOG\'!A:A)))'
    ws['G15'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=1,"",INDEX(\'SESSION LOG\'!J:J,COUNTA(\'SESSION LOG\'!A:A)))'
    
    ws['B16'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=2,"",INDEX(\'SESSION LOG\'!A:A,COUNTA(\'SESSION LOG\'!A:A)-1))'
    ws['C16'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=2,"",INDEX(\'SESSION LOG\'!B:B,COUNTA(\'SESSION LOG\'!A:A)-1))'
    ws['D16'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=2,"",INDEX(\'SESSION LOG\'!C:C,COUNTA(\'SESSION LOG\'!A:A)-1))'
    ws['E16'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=2,"",INDEX(\'SESSION LOG\'!F:F,COUNTA(\'SESSION LOG\'!A:A)-1))'
    ws['F16'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=2,"",INDEX(\'SESSION LOG\'!G:G,COUNTA(\'SESSION LOG\'!A:A)-1))'
    ws['G16'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=2,"",INDEX(\'SESSION LOG\'!J:J,COUNTA(\'SESSION LOG\'!A:A)-1))'
    
    ws['B17'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=3,"",INDEX(\'SESSION LOG\'!A:A,COUNTA(\'SESSION LOG\'!A:A)-2))'
    ws['C17'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=3,"",INDEX(\'SESSION LOG\'!B:B,COUNTA(\'SESSION LOG\'!A:A)-2))'
    ws['D17'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=3,"",INDEX(\'SESSION LOG\'!C:C,COUNTA(\'SESSION LOG\'!A:A)-2))'
    ws['E17'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=3,"",INDEX(\'SESSION LOG\'!F:F,COUNTA(\'SESSION LOG\'!A:A)-2))'
    ws['F17'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=3,"",INDEX(\'SESSION LOG\'!G:G,COUNTA(\'SESSION LOG\'!A:A)-2))'
    ws['G17'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=3,"",INDEX(\'SESSION LOG\'!J:J,COUNTA(\'SESSION LOG\'!A:A)-2))'
    
    ws['A19'] = 'WEEKLY REPORT (Auto-Generated - Copy & Paste)'
    style_section_header(ws['A19'])
    ws.merge_cells('A19:G19')
    
    ws['A20'] = 'Report Text'
    ws.merge_cells('B20:G25')
    ws['B20'] = '="REFINEMENT PROGRESS - Week of "&TEXT(TODAY(),"MMM DD, YYYY")&CHAR(10)&CHAR(10)&"✅ STORIES REFINED THIS WEEK"&CHAR(10)&"   • Sessions conducted: "&COUNTIFS(\'SESSION LOG\'!C:C,">="&TODAY()-7,\'SESSION LOG\'!C:C,"<="&TODAY())&CHAR(10)&"   • Stories moved to Ready: "&SUMIFS(\'SESSION LOG\'!F:F,\'SESSION LOG\'!C:C,">="&TODAY()-7,\'SESSION LOG\'!C:C,"<="&TODAY())&CHAR(10)&"   • Cumulative progress: "&C9&"/84 ("&ROUND(C11,0)&"%)"&CHAR(10)&CHAR(10)&"📈 VELOCITY METRICS"&CHAR(10)&"   • Current velocity: "&ROUND(F9,1)&" stories/session"&CHAR(10)&"   • Target velocity: 6.5 stories/session"&CHAR(10)&"   • Status: "&F6&CHAR(10)&CHAR(10)&"📅 TIMELINE"&CHAR(10)&"   • Sessions remaining: "&(15-COUNTA(\'SESSION LOG\'!F2:F16))&CHAR(10)&"   • Projected completion: "&F11&CHAR(10)&"   • On track for Sprint 5: "&IF(F11="Sprint 5","YES ✅",IF(F11="Sprint 6","ALMOST ⚠️","NO ❌"))'
    ws['B20'].alignment = Alignment(wrap_text=True, vertical='top')
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 12

def create_session_log_sheet(wb):
    """Create SESSION LOG sheet"""
    ws = wb.create_sheet('SESSION LOG')
    
    ws['A1'] = 'SESSION LOG - Data Entry (Update After Each Session)'
    ws['A1'].font = Font(bold=True, size=12, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws.merge_cells('A1:K1')
    
    headers = ['Session #', 'Sprint', 'Date', 'Day', 'Target Stories', 'Stories Refined', 
               'Cumulative', 'Remaining', 'Velocity', 'Status', 'Notes/Blockers']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        style_header(cell)
    
    sessions = [
        (1, 1, '2025-11-05', 'Tuesday', 7, 0, 'Learning session'),
        (2, 1, '2025-11-07', 'Thursday', 7, 1, 'Team alignment issues'),
        (3, 1, '2025-11-12', 'Tuesday', 7, '', ''),
        (4, 2, '2025-11-19', 'Tuesday', 7, '', ''),
        (5, 2, '2025-11-21', 'Thursday', 7, '', ''),
        (6, 2, '2025-11-26', 'Tuesday', 7, '', ''),
        (7, 3, '2025-12-03', 'Tuesday', 7, '', ''),
        (8, 3, '2025-12-05', 'Thursday', 7, '', ''),
        (9, 3, '2025-12-10', 'Tuesday', 7, '', ''),
        (10, 4, '2025-12-17', 'Tuesday', 7, '', ''),
        (11, 4, '2025-12-19', 'Thursday', 7, '', ''),
        (12, 4, '2025-12-24', 'Tuesday', 7, '', ''),
        (13, 5, '2026-01-07', 'Tuesday', 7, '', ''),
        (14, 5, '2026-01-09', 'Thursday', 7, '', ''),
        (15, 5, '2026-01-14', 'Tuesday', 7, '', ''),
    ]
    
    for idx, (session, sprint, date, day, target, refined, notes) in enumerate(sessions, start=3):
        row = idx
        ws[f'A{row}'] = session
        ws[f'B{row}'] = sprint
        ws[f'C{row}'] = date
        ws[f'D{row}'] = day
        ws[f'E{row}'] = target
        if refined != '':
            ws[f'F{row}'] = refined
        ws[f'G{row}'] = f'=SUM($F$3:F{row})'
        ws[f'H{row}'] = f'=84-G{row}'
        ws[f'I{row}'] = f'=F{row}'
        ws[f'J{row}'] = f'=IF(ISBLANK(I{row}),"",IF(I{row}>=6.5,"🟢",IF(I{row}>=4,"🟡","🔴")))'
        ws[f'K{row}'] = notes
    
    ws['A19'] = 'INSTRUCTIONS:'
    ws['A19'].font = Font(bold=True, color='FF0000')
    ws['A20'] = 'After each refinement session, update ONLY column F (Stories Refined) with the number of stories moved to "Ready"'
    ws.merge_cells('A20:K20')
    ws['A21'] = 'Everything else calculates automatically!'
    ws['A21'].font = Font(bold=True, color='00B050')
    
    for col, width in zip(['A','B','C','D','E','F','G','H','I','J','K'], [10,8,12,10,14,14,12,12,10,8,30]):
        ws.column_dimensions[col].width = width

def create_burndown_sheet(wb):
    """Create BURNDOWN CHART DATA sheet"""
    ws = wb.create_sheet('BURNDOWN CHART DATA')
    
    ws['A1'] = 'BURNDOWN CHART DATA'
    ws['A1'].font = Font(bold=True, size=12, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws.merge_cells('A1:F1')
    
    headers = ['Session #', 'Sprint', 'Stories Remaining (Ideal)', 'Stories Remaining (Actual)', 'Variance', 'Status']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        style_header(cell)
    
    ws['A3'] = 0
    ws['B3'] = 1
    ws['C3'] = 84
    ws['D3'] = 84
    ws['E3'] = 0
    ws['F3'] = 'Start'
    
    for row in range(4, 19):
        session_row = row - 1
        ws[f'A{row}'] = f'=INDEX(\'SESSION LOG\'!A:A,{session_row})'
        ws[f'B{row}'] = f'=INDEX(\'SESSION LOG\'!B:B,{session_row})'
        ws[f'C{row}'] = f'=84-5.6*({row-3})'
        ws[f'D{row}'] = f'=INDEX(\'SESSION LOG\'!H:H,{session_row})'
        ws[f'E{row}'] = f'=D{row}-C{row}'
        ws[f'F{row}'] = f'=IF(E{row}>=0,"🟢",IF(E{row}>=-5,"🟡","🔴"))'
    
    ws['A21'] = 'Create Line Chart: X-axis = Session # | Y-axis = Stories Remaining Ideal (C) and Actual (D)'
    ws['A21'].font = Font(italic=True)
    ws.merge_cells('A21:F21')
    
    for col in ['A','B','C','D','E','F']:
        ws.column_dimensions[col].width = 18

def create_burnup_sheet(wb):
    """Create BURNUP CHART DATA sheet"""
    ws = wb.create_sheet('BURNUP CHART DATA')
    
    ws['A1'] = 'BURNUP CHART DATA'
    ws['A1'].font = Font(bold=True, size=12, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws.merge_cells('A1:F1')
    
    headers = ['Session #', 'Sprint', 'Stories Refined (Target)', 'Stories Refined (Actual)', 'Total Scope', 'Gap']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        style_header(cell)
    
    ws['A3'] = 0
    ws['B3'] = 1
    ws['C3'] = 0
    ws['D3'] = 0
    ws['E3'] = 84
    ws['F3'] = 0
    
    for row in range(4, 19):
        session_row = row - 1
        ws[f'A{row}'] = f'=INDEX(\'SESSION LOG\'!A:A,{session_row})'
        ws[f'B{row}'] = f'=INDEX(\'SESSION LOG\'!B:B,{session_row})'
        ws[f'C{row}'] = f'=5.6*({row-3})'
        ws[f'D{row}'] = f'=INDEX(\'SESSION LOG\'!G:G,{session_row})'
        ws[f'E{row}'] = 84
        ws[f'F{row}'] = f'=D{row}-C{row}'
    
    ws['A21'] = 'Create Line Chart: X-axis = Session # | Y-axis = Target (C), Actual (D), and Total Scope (E)'
    ws['A21'].font = Font(italic=True)
    ws.merge_cells('A21:F21')
    
    for col in ['A','B','C','D','E','F']:
        ws.column_dimensions[col].width = 18

def create_epic_progress_sheet(wb):
    """Create EPIC PROGRESS sheet"""
    ws = wb.create_sheet('EPIC PROGRESS')
    
    ws['A1'] = 'EPIC PROGRESS'
    ws['A1'].font = Font(bold=True, size=12, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws.merge_cells('A1:I1')
    
    headers = ['Epic ID', 'Epic Name', 'Total Stories', 'Open', 'Ready', 'Other', '% Complete', 'Priority', 'Sprint Target']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        style_header(cell)
    
    epics = [
        ('SSF-123', 'Fishing Operations', 14, 13, 1, 0, 'HIGH', '1-2'),
        ('SSF-96', 'Login & Session', 12, 6, 6, 0, 'CRITICAL', '1'),
        ('SSF-122', 'Departure & Catches', 10, 10, 0, 0, 'HIGH', '2-3'),
        ('SSF-93', 'Gear Management', 10, 9, 1, 0, 'MEDIUM', '2-3'),
        ('SSF-95', 'Map & Location', 8, 6, 2, 0, 'MEDIUM', '3-4'),
        ('SSF-84', 'Technical Setup', 8, 3, 5, 0, 'HIGH', '1'),
        ('SSF-165', 'Backend Infrastructure', 3, 3, 0, 0, 'MEDIUM', '4-5'),
        ('SSF-166', 'Internationalization', 4, 3, 1, 0, 'LOW', '5'),
        ('SSF-162', 'Menu & Navigation', 3, 2, 1, 0, 'MEDIUM', '3'),
        ('SSF-94', 'Home & Widgets', 5, 4, 1, 0, 'MEDIUM', '2'),
        ('SSF-86', 'Tracking', 6, 1, 5, 0, 'HIGH', '1'),
        ('SSF-97', 'CI/CD & Deployment', 5, 5, 0, 0, 'MEDIUM', '4-5'),
        ('SSF-132', 'FMC Integration', 1, 1, 0, 0, 'MEDIUM', '4'),
    ]
    
    for idx, (epic_id, name, total, open_stories, ready, other, priority, sprint) in enumerate(epics, start=3):
        row = idx
        ws[f'A{row}'] = epic_id
        ws[f'B{row}'] = name
        ws[f'C{row}'] = total
        ws[f'D{row}'] = open_stories
        ws[f'E{row}'] = ready
        ws[f'F{row}'] = other
        ws[f'G{row}'] = f'=IF(C{row}=0,0,E{row}/C{row}*100)'
        ws[f'H{row}'] = priority
        ws[f'I{row}'] = sprint
    
    ws['A17'] = 'Create Bar Chart showing % Complete by Epic'
    ws['A17'].font = Font(italic=True)
    ws.merge_cells('A17:I17')
    
    for col, width in zip(['A','B','C','D','E','F','G','H','I'], [12,25,12,8,8,8,12,12,14]):
        ws.column_dimensions[col].width = width

def create_session_calendar_sheet(wb):
    """Create SESSION CALENDAR sheet"""
    ws = wb.create_sheet('SESSION CALENDAR')
    
    ws['A1'] = 'SESSION CALENDAR - All 13 Remaining Sessions'
    ws['A1'].font = Font(bold=True, size=12, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws.merge_cells('A1:K1')
    
    headers = ['Session #', 'Sprint', 'Week', 'Date', 'Day', 'Time', 'Duration (min)', 
               'Target Stories', 'Pre-Prep Deadline', 'Status', 'Notes']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col)
        cell.value = header
        style_header(cell)
    
    sessions = [
        (3, 1, 2, '2025-11-12', 'Tuesday', '14:00', 60, 7, '2025-11-10', '🟡 This Week', 'Focus on quick wins'),
        (4, 2, 1, '2025-11-19', 'Tuesday', '14:00', 60, 7, '2025-11-17', '🟢 Upcoming', 'Gear Management Backend'),
        (5, 2, 1, '2025-11-21', 'Thursday', '14:00', 60, 7, '2025-11-19', '🟢 Upcoming', 'Gear Management Frontend'),
        (6, 2, 2, '2025-11-26', 'Tuesday', '14:00', 60, 7, '2025-11-24', '🟢 Upcoming', 'Map & Location'),
        (7, 3, 1, '2025-12-03', 'Tuesday', '14:00', 60, 7, '2025-12-01', '🟢 Upcoming', 'Departure Operations'),
        (8, 3, 1, '2025-12-05', 'Thursday', '14:00', 60, 7, '2025-12-03', '🟢 Upcoming', 'Catches & Preferences'),
        (9, 3, 2, '2025-12-10', 'Tuesday', '14:00', 60, 7, '2025-12-08', '🟢 Upcoming', 'Backend & Forms'),
        (10, 4, 1, '2025-12-17', 'Tuesday', '14:00', 60, 7, '2025-12-15', '🟢 Upcoming', 'Authentication'),
        (11, 4, 1, '2025-12-19', 'Thursday', '14:00', 60, 7, '2025-12-17', '🟢 Upcoming', 'Infrastructure'),
        (12, 4, 2, '2025-12-24', 'Tuesday', '14:00', 60, 7, '2025-12-22', '🟢 Upcoming', 'Buffer/Cleanup'),
        (13, 5, 1, '2026-01-07', 'Tuesday', '14:00', 60, 7, '2026-01-05', '🟢 Upcoming', 'Final push'),
        (14, 5, 1, '2026-01-09', 'Thursday', '14:00', 60, 7, '2026-01-07', '🟢 Upcoming', 'Final push'),
        (15, 5, 2, '2026-01-14', 'Tuesday', '14:00', 60, 7, '2026-01-12', '🟢 Upcoming', 'Completion'),
    ]
    
    for idx, session_data in enumerate(sessions, start=3):
        row = idx
        for col, value in enumerate(session_data, start=1):
            ws.cell(row=row, column=col, value=value)
    
    ws['A17'] = 'INSTRUCTIONS:'
    ws['A17'].font = Font(bold=True, color='FF0000')
    ws['A18'] = 'Adjust dates if needed based on your actual sprint schedule'
    ws['A19'] = 'Status will update as sessions complete (manually change from 🟢 to ⚪)'
    
    widths = [10, 8, 6, 12, 10, 8, 14, 14, 18, 14, 30]
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width

def create_quick_reference_sheet(wb):
    """Create QUICK REFERENCE sheet with session agenda"""
    ws = wb.create_sheet('QUICK REFERENCE')
    
    ws['A1'] = '60-MINUTE SESSION AGENDA TEMPLATE'
    ws['A1'].font = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws.merge_cells('A1:D1')
    
    ws['A3'] = 'TIME'
    ws['B3'] = 'ACTIVITY'
    ws['C3'] = 'DURATION'
    ws['D3'] = 'DETAILS'
    for col in ['A3', 'B3', 'C3', 'D3']:
        style_header(ws[col])
    
    agenda = [
        ('00:00-05:00', 'RECAP & ADJUST', '5 min', 'Previous session results, adjustments needed'),
        ('05:00-10:00', 'SILENT READING TIME', '5 min', 'Team reviews stories individually - NO DISCUSSION'),
        ('10:00-50:00', 'STORY REFINEMENT', '40 min', '7 minutes per story, strict time-boxing'),
        ('50:00-58:00', 'WRAP-UP', '8 min', 'Count refined stories, update tracker, plan next'),
        ('58:00-60:00', 'BUFFER', '2 min', 'Overflow or early finish'),
    ]
    
    for idx, (time, activity, duration, details) in enumerate(agenda, start=4):
        ws[f'A{idx}'] = time
        ws[f'B{idx}'] = activity
        ws[f'C{idx}'] = duration
        ws[f'D{idx}'] = details
    
    ws['A10'] = 'STORY REFINEMENT CHECKLIST (Use for each story):'
    ws['A10'].font = Font(bold=True)
    ws.merge_cells('A10:D10')
    
    checklist = [
        'Business value understood?',
        'Business rules clarified?',
        'Acceptance criteria agreed?',
        'Technical approach confirmed?',
        'Dependencies identified?',
        'Edge cases discussed?',
        'Estimation consensus?',
        'Ready to move to Ready status?',
    ]
    
    for idx, item in enumerate(checklist, start=11):
        ws[f'A{idx}'] = item
        ws[f'B{idx}'] = '☐'
    
    ws['A20'] = 'SUCCESS CRITERIA'
    ws['A20'].font = Font(bold=True, color='FFFFFF')
    ws['A20'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    ws.merge_cells('A20:D20')
    
    ws['A21'] = 'Target: 6-7 stories to "Ready"'
    ws['A22']
