#!/usr/bin/env python3
"""
Convert Gear Registration Analysis CSV to XLSX format
"""
import csv
import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("Error: openpyxl library not found.")
    print("Please install it with: pip install openpyxl")
    sys.exit(1)

# File paths
csv_file = Path(__file__).parent / "Gear_Registration_Analysis.csv"
xlsx_file = Path(__file__).parent / "Gear_Registration_Analysis.xlsx"

# Create workbook
wb = Workbook()
ws = wb.active
ws.title = "Gear Registration Analysis"

# Read CSV and write to Excel
with open(csv_file, 'r', encoding='utf-8') as f:
    reader = csv.reader(f)
    for row_idx, row in enumerate(reader, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Format header row
            if row_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            else:
                # Format data rows
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Color code complexity ratings
                if col_idx == 11:  # Complexity Rating column
                    if value == "Simple":
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                    elif value == "Medium":
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                    elif value == "Complex":
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                
                # Add borders
                thin_border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
                cell.border = thin_border

# Auto-adjust column widths
column_widths = {
    'A': 18,  # Category
    'B': 18,  # Subcategory
    'C': 30,  # Gear Type Name
    'D': 8,   # Code
    'E': 35,  # Required Characteristics
    'F': 45,  # Characteristic 1
    'G': 45,  # Characteristic 2
    'H': 45,  # Characteristic 3
    'I': 35,  # Registration Approach
    'J': 60,  # Key Questions
    'K': 12,  # Complexity Rating
    'L': 50,  # Implementation Notes
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Set row height for header
ws.row_dimensions[1].height = 30

# Freeze header row
ws.freeze_panes = "A2"

# Add auto-filter
ws.auto_filter.ref = ws.dimensions

# Save workbook
wb.save(xlsx_file)
print(f"✓ Successfully created: {xlsx_file}")
print(f"  - {ws.max_row - 1} gear types analyzed")
print(f"  - {ws.max_column} columns of data")
print(f"  - Header row frozen for easy scrolling")
print(f"  - Auto-filter enabled")
print(f"  - Complexity ratings color-coded (Green=Simple, Yellow=Medium, Red=Complex)")
