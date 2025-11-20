#!/usr/bin/env python3
"""
Fix #NAME? errors in EPIC PROGRESS sheet by ensuring proper formula formatting
"""

import openpyxl
from openpyxl.utils import get_column_letter

def fix_epic_progress_formulas(tracker_file):
    """Fix formulas in EPIC PROGRESS to eliminate #NAME? errors"""
    
    print("🔧 Fixing EPIC PROGRESS formulas...")
    
    wb = openpyxl.load_workbook(tracker_file)
    ws_epic = wb['EPIC PROGRESS']
    ws_jira = wb['JIRA DATA']
    
    # Get list of unique epic names from JIRA DATA
    epic_names = set()
    for row in ws_jira.iter_rows(min_row=3, values_only=True):
        if row[5]:  # Column F - Epic Link
            epic_names.add(row[5])
    
    epic_list = sorted(list(epic_names))
    
    print(f"Found {len(epic_list)} unique epics")
    
    # Clear existing formulas starting from row 3
    if ws_epic.max_row > 2:
        ws_epic.delete_rows(3, ws_epic.max_row - 2)
    
    # Write epic names and formulas
    for idx, epic_name in enumerate(epic_list, start=3):
        print(f"  Processing: {epic_name}")
        
        # Column A & B: Epic Name
        ws_epic.cell(row=idx, column=1).value = epic_name
        ws_epic.cell(row=idx, column=2).value = epic_name
        
        # Column C: Total (using COUNTIF with proper sheet reference)
        ws_epic.cell(row=idx, column=3).value = f"=COUNTIF('JIRA DATA'!F:F,A{idx})"
        
        # Column D: Open
        ws_epic.cell(row=idx, column=4).value = f"=COUNTIFS('JIRA DATA'!F:F,A{idx},'JIRA DATA'!E:E,\"Open\")"
        
        # Column E: Ready
        ws_epic.cell(row=idx, column=5).value = f"=COUNTIFS('JIRA DATA'!F:F,A{idx},'JIRA DATA'!E:E,\"Ready\")"
        
        # Column F: Other
        ws_epic.cell(row=idx, column=6).value = f"=C{idx}-D{idx}-E{idx}"
        
        # Column G: % Complete
        ws_epic.cell(row=idx, column=7).value = f"=IF(C{idx}=0,0,E{idx}/C{idx})"
        ws_epic.cell(row=idx, column=7).number_format = '0.0%'
        
        # Columns H-I: Manual entry (Priority, Sprint)
        
        # Column J: PI Filter (manual entry)
        
        # Column K: Refined in PI
        ws_epic.cell(row=idx, column=11).value = f"=IF(J{idx}=\"\",\"-\",COUNTIFS('JIRA DATA'!F:F,A{idx},'JIRA DATA'!J:J,J{idx},'JIRA DATA'!E:E,\"Ready\"))"
        
        # Column L: Total in PI
        ws_epic.cell(row=idx, column=12).value = f"=IF(J{idx}=\"\",\"-\",COUNTIFS('JIRA DATA'!F:F,A{idx},'JIRA DATA'!J:J,J{idx}))"
        
        # Column M: PI Progress
        ws_epic.cell(row=idx, column=13).value = f"=IF(J{idx}=\"\",\"-\",K{idx}&\"/\"&L{idx})"
    
    # Force calculation mode to automatic
    wb.calculation.calcMode = 'auto'
    
    print(f"\n✅ Updated formulas for {len(epic_list)} epics")
    print(f"💾 Saving workbook...")
    
    wb.save(tracker_file)
    wb.close()
    
    print("\n" + "="*60)
    print("✅ FORMULAS FIXED!")
    print("="*60)
    print("\n📝 Next Steps:")
    print("  1. Open the Excel file")
    print("  2. If you still see #NAME? errors:")
    print("     - Press Ctrl+Alt+F9 to force recalculation")
    print("     - Or go to Formulas > Calculate Now")
    print("  3. Save the file in Excel")
    print("\n⚠️  Note: Sometimes Excel needs to recalculate formulas")
    print("    after they're written programmatically.")

if __name__ == '__main__':
    tracker_file = 'DG_MARE_TEAMS_SSF_Refinement_Tracker.xlsx'
    fix_epic_progress_formulas(tracker_file)
