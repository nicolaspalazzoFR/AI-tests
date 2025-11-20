#!/usr/bin/env python3
"""
Create Microsoft 365 Compatible Tracker
- All features from previous scripts
- NO chart objects (these cause Office 365 online corruption)
- Fully editable in Microsoft 365 online
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

def extract_pi_from_labels(labels_str):
    """Extract PI label from Labels string"""
    if not labels_str:
        return ""
    match = re.search(r'SSF_PI#\d+', str(labels_str))
    return match.group(0) if match else ""

def load_jira_data(source_file):
    """Load data from SSF_BACKLOG_EXPORT_10_11.xlsx"""
    wb_source = openpyxl.load_workbook(source_file, data_only=True)
    ws_source = wb_source[wb_source.sheetnames[0]]
    
    data = []
    for row in ws_source.iter_rows(min_row=5, values_only=True):
        if row[1]:  # if Issue key exists
            data.append({
                'Issue Type': row[0],
                'issueKey': row[1],
                'Issue id': row[2] if len(row) > 2 else None,
                'Summary': row[3],
                'Status': row[4],
                'Epic Link': row[5],
                'Story Points': row[6],
                'Labels': row[7] if len(row) > 7 else "",
                'Sprint': row[8] if len(row) > 8 else ""
            })
    
    wb_source.close()
    return data

def update_jira_data_sheet(ws, data):
    """Update JIRA DATA sheet with Labels and PI columns"""
    print("\n📝 Updating JIRA DATA sheet...")
    
    headers = ['Issue Type', 'Issue key', 'Issue id', 'Summary', 'Status', 
               'Epic Link', 'Story Points', 'Labels', 'Sprint', 'PI']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.delete_rows(3, ws.max_row)
    
    for row_idx, item in enumerate(data, start=3):
        ws.cell(row=row_idx, column=1, value=item['Issue Type'])
        ws.cell(row=row_idx, column=2, value=item['issueKey'])
        ws.cell(row=row_idx, column=3, value=item['Issue id'])
        ws.cell(row=row_idx, column=4, value=item['Summary'])
        ws.cell(row=row_idx, column=5, value=item['Status'])
        ws.cell(row=row_idx, column=6, value=item['Epic Link'])
        ws.cell(row=row_idx, column=7, value=item['Story Points'])
        ws.cell(row=row_idx, column=8, value=item['Labels'])
        ws.cell(row=row_idx, column=9, value=item['Sprint'])
        ws.cell(row=row_idx, column=10, value=extract_pi_from_labels(item['Labels']))
    
    print(f"   ✅ Imported {len(data)} issues")

def create_pi_progress_sheet(wb, ws_jira_data):
    """Create PI PROGRESS sheet"""
    print("\n📊 Creating PI PROGRESS sheet...")
    
    if 'PI PROGRESS' in wb.sheetnames:
        del wb['PI PROGRESS']
    
    ws = wb.create_sheet('PI PROGRESS', 2)
    
    ws['A1'] = 'PI PROGRESS TRACKER'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ['PI Label', 'Total Stories', 'Refined (Ready)', 'To Refine (Open)', 
               'Other Status', '% Complete', 'Progress', 'Status']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    pi_list = set()
    for row in ws_jira_data.iter_rows(min_row=3, values_only=True):
        pi_val = row[9] if len(row) > 9 else None
        if pi_val and pi_val.startswith('SSF_PI#'):
            pi_list.add(pi_val)
    
    pi_list = sorted(list(pi_list))
    
    for idx, pi in enumerate(pi_list, start=3):
        ws.cell(row=idx, column=1, value=pi)
        ws.cell(row=idx, column=2, value=f'=COUNTIF(\'JIRA DATA\'!J:J,A{idx})')
        ws.cell(row=idx, column=3, value=f'=COUNTIFS(\'JIRA DATA\'!J:J,A{idx},\'JIRA DATA\'!E:E,"Ready")')
        ws.cell(row=idx, column=4, value=f'=COUNTIFS(\'JIRA DATA\'!J:J,A{idx},\'JIRA DATA\'!E:E,"Open")')
        ws.cell(row=idx, column=5, value=f'=B{idx}-C{idx}-D{idx}')
        ws.cell(row=idx, column=6, value=f'=IF(B{idx}=0,0,C{idx}/B{idx})')
        ws.cell(row=idx, column=6).number_format = '0.0%'
        ws.cell(row=idx, column=7, value=f'=C{idx}&"/"&B{idx}')
        ws.cell(row=idx, column=8, value=f'=IF(F{idx}>0.7,"🟢",IF(F{idx}>0.3,"🟡","🔴"))')
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col].width = 15
    
    print(f"   ✅ Created with {len(pi_list)} PIs")

def fix_epic_progress_sheet(ws, ws_jira_data):
    """Fix EPIC PROGRESS with epic names"""
    print("\n🔧 Fixing EPIC PROGRESS...")
    
    epic_names = set()
    for row in ws_jira_data.iter_rows(min_row=3, values_only=True):
        epic = row[5]
        if epic and epic != "Epic Link":
            epic_names.add(epic)
    
    epic_list = sorted(list(epic_names))
    
    if ws.max_row > 2:
        ws.delete_rows(3, ws.max_row - 2)
    
    for idx, epic_name in enumerate(epic_list, start=3):
        ws.cell(row=idx, column=1, value=epic_name)
        ws.cell(row=idx, column=2, value=epic_name)
        ws.cell(row=idx, column=3, value=f"=COUNTIF('JIRA DATA'!F:F,A{idx})")
        ws.cell(row=idx, column=4, value=f"=COUNTIFS('JIRA DATA'!F:F,A{idx},'JIRA DATA'!E:E,\"Open\")")
        ws.cell(row=idx, column=5, value=f"=COUNTIFS('JIRA DATA'!F:F,A{idx},'JIRA DATA'!E:E,\"Ready\")")
        ws.cell(row=idx, column=6, value=f"=C{idx}-D{idx}-E{idx}")
        ws.cell(row=idx, column=7, value=f"=IF(C{idx}=0,0,E{idx}/C{idx})")
        ws.cell(row=idx, column=7).number_format = '0.0%'
        
        start_col = ws.max_column - 2
        col_j = get_column_letter(start_col)
        col_k = get_column_letter(start_col + 1)
        col_l = get_column_letter(start_col + 2)
        col_m = get_column_letter(start_col + 3)
        
        ws.cell(row=idx, column=start_col + 1, 
               value=f"=IF({col_j}{idx}=\"\",\"-\",COUNTIFS('JIRA DATA'!F:F,A{idx},'JIRA DATA'!J:J,{col_j}{idx},'JIRA DATA'!E:E,\"Ready\"))")
        ws.cell(row=idx, column=start_col + 2,
               value=f"=IF({col_j}{idx}=\"\",\"-\",COUNTIFS('JIRA DATA'!F:F,A{idx},'JIRA DATA'!J:J,{col_j}{idx}))")
        ws.cell(row=idx, column=start_col + 3,
               value=f"=IF({col_j}{idx}=\"\",\"-\",{col_k}{idx}&\"/\"&{col_l}{idx})")
    
    print(f"   ✅ Fixed {len(epic_list)} epics")

def add_pi_column_to_session_log(ws_session):
    """Add PI(s) Worked column"""
    print("\n🔧 Updating SESSION LOG...")
    
    ws_session.cell(row=2, column=11, value="PI(s) Worked")
    ws_session.cell(row=2, column=11).font = Font(bold=True, color="FFFFFF")
    ws_session.cell(row=2, column=11).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws_session.cell(row=2, column=11).alignment = Alignment(horizontal='center', vertical='center')
    
    ws_session.cell(row=3, column=11, value="Enter PI labels: SSF_PI#01, SSF_PI#02, etc.")
    ws_session.cell(row=3, column=11).font = Font(italic=True, color="999999")
    ws_session.column_dimensions['K'].width = 25
    
    print("   ✅ Added PI column")

def create_pi_session_log(wb, pi_label, ws_jira_data):
    """Create PI SESSION LOG (no charts)"""
    sheet_name = f"{pi_label} SESSION LOG"
    
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name)
    
    ws['A1'] = f'{pi_label} REFINEMENT SESSION LOG'
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:I1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    headers = ['Session #', 'Sprint', 'Date', 'PI(s) Worked', 'Target Stories', 
               'Stories Refined', 'Cumulative', 'Remaining', 'Velocity']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    pi_total_formula = f'=COUNTIF(\'JIRA DATA\'!J:J,"{pi_label}")'
    
    for row_idx in range(3, 20):
        ws.cell(row=row_idx, column=1, value=row_idx - 2)
        ws.cell(row=row_idx, column=2, value=f"='SESSION LOG'!B{row_idx}")
        ws.cell(row=row_idx, column=3, value=f"='SESSION LOG'!C{row_idx}")
        ws.cell(row=row_idx, column=4, value=f"='SESSION LOG'!K{row_idx}")
        ws.cell(row=row_idx, column=5, value=f"='SESSION LOG'!E{row_idx}")
        ws.cell(row=row_idx, column=6, 
               value=f'=IF(ISNUMBER(SEARCH("{pi_label}",\'SESSION LOG\'!K{row_idx})),\'SESSION LOG\'!F{row_idx},0)')
        ws.cell(row=row_idx, column=7, value=f"=SUM($F$3:F{row_idx})")
        ws.cell(row=row_idx, column=8, value=f"={pi_total_formula}-G{row_idx}")
        
        if row_idx == 3:
            ws.cell(row=row_idx, column=9, value=f"=F{row_idx}")
        else:
            ws.cell(row=row_idx, column=9, value=f"=AVERAGE($F$3:F{row_idx})")
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[col].width = 15

def create_pi_burndown_data(wb, pi_label):
    """Create PI BURNDOWN DATA sheet (NO CHART OBJECT)"""
    sheet_name = f"{pi_label} BURNDOWN"
    
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name)
    
    ws['A1'] = 'TARGET SESSIONS:'
    ws['A1'].font = Font(bold=True)
    ws['B1'] = 15
    ws['B1'].font = Font(bold=True, color="FF0000")
    ws['B1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    ws['A2'] = 'TOTAL STORIES:'
    ws['A2'].font = Font(bold=True)
    ws['B2'] = f'=COUNTIF(\'JIRA DATA\'!J:J,"{pi_label}")'
    ws['B2'].font = Font(bold=True)
    
    ws['D1'] = f'{pi_label} BURNDOWN DATA'
    ws['D1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['D1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('D1:H1')
    ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
    
    ws['A4'] = 'Session #'
    ws['B4'] = 'Sprint'
    ws['C4'] = 'IDEAL Remaining'
    ws['D4'] = 'ACTUAL Remaining'
    ws['E4'] = 'Variance'
    
    for cell in [ws['A4'], ws['B4'], ws['C4'], ws['D4'], ws['E4']]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    session_log_sheet = f"{pi_label} SESSION LOG"
    
    for row_idx in range(5, 22):
        ws.cell(row=row_idx, column=1, value=row_idx - 5)
        ws.cell(row=row_idx, column=2, value=f"='{session_log_sheet}'!B{row_idx-2}")
        ws.cell(row=row_idx, column=3, value=f"=$B$2-($B$2/$B$1)*A{row_idx}")
        ws.cell(row=row_idx, column=4, value=f"='{session_log_sheet}'!H{row_idx-2}")
        ws.cell(row=row_idx, column=5, value=f"=D{row_idx}-C{row_idx}")
    
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 15
    
    # Add instruction for manual chart creation
    ws['G4'] = 'TO CREATE CHART:'
    ws['G4'].font = Font(bold=True, color="FF0000")
    ws['G5'] = '1. Select cells C4:D21'
    ws['G6'] = '2. Insert > Chart > Line'
    ws['G7'] = '3. Title: See cell D1'
    
    print(f"   ✅ Created {sheet_name} (data only)")

def update_dashboard(ws):
    """Update DASHBOARD"""
    print("\n📋 Updating DASHBOARD...")
    
    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1, value="📊 PI PROGRESS")
    ws.cell(row=note_row, column=1).font = Font(size=14, bold=True, color="4472C4")
    ws.cell(row=note_row + 1, column=1, value="See PI PROGRESS sheet for PI tracking →")
    ws.cell(row=note_row + 1, column=1).font = Font(italic=True)
    
    print("   ✅ Updated")

def main():
    source_file = 'SSF_BACKLOG_EXPORT_10_11.xlsx'
    tracker_file = 'DG_MARE_TEAMS_SSF_Refinement_Tracker.xlsx'
    
    print("🚀 Creating Microsoft 365 Compatible Tracker")
    print("   (Without chart objects)")
    print("="*60)
    
    print(f"\n📂 Loading JIRA data from {source_file}...")
    data = load_jira_data(source_file)
    print(f"   Found {len(data)} issues")
    
    print(f"\n📂 Loading tracker: {tracker_file}")
    wb = openpyxl.load_workbook(tracker_file)
    
    ws_jira_data = wb['JIRA DATA']
    update_jira_data_sheet(ws_jira_data, data)
    
    create_pi_progress_sheet(wb, ws_jira_data)
    
    if 'EPIC PROGRESS' in wb.sheetnames:
        fix_epic_progress_sheet(wb['EPIC PROGRESS'], ws_jira_data)
    
    if 'SESSION LOG' in wb.sheetnames:
        add_pi_column_to_session_log(wb['SESSION LOG'])
    
    print("\n📈 Creating PI tracking sheets...")
    pi_list = ['SSF_PI#01', 'SSF_PI#02']
    
    for pi in pi_list:
        print(f"\n   Processing {pi}...")
        create_pi_session_log(wb, pi, ws_jira_data)
        create_pi_burndown_data(wb, pi)
    
    if 'DASHBOARD' in wb.sheetnames:
        update_dashboard(wb['DASHBOARD'])
    
    # Remove any existing problematic sheets
    for sheet_name in ['SSF_PI#01 BURNDOWN OLD', 'SSF_PI#02 BURNDOWN OLD']:
        if sheet_name in wb.sheetnames:
            del wb[sheet_name]
    
    # Set calculation mode to automatic
    wb.calculation.calcMode = 'auto'
    
    print(f"\n💾 Saving to {tracker_file}...")
    wb.save(tracker_file)
    wb.close()
    
    print("\n" + "="*60)
    print("✅ MICROSOFT 365 COMPATIBLE TRACKER CREATED!")
    print("="*60)
    print("\n📊 What's Included:")
    print("   ✅ JIRA DATA with Labels + PI columns")
    print("   ✅ PI PROGRESS sheet")
    print("   ✅ EPIC PROGRESS fixed (epic names)")
    print("   ✅ SESSION LOG with PI(s) Worked column")
    print("   ✅ PI SESSION LOG sheets (filtered)")
    print("   ✅ PI BURNDOWN DATA sheets")
    print("   ✅ All formulas working")
    print("   ❌ NO chart objects (Office 365 compatible)")
    
    print("\n📝 To Use:")
    print("   1. Upload to Office 365 online")
    print("   2. File will open and be editable")
    print("   3. Use PI BURNDOWN data tables to track")
    print("   4. Optionally create charts manually:")
    print("      - Go to PI BURNDOWN sheet")
    print("      - Select cells C4:D21")
    print("      - Insert > Chart > Line")
    
    print("\n🎯 Key Features:")
    print("   • Cell B1 in burndown sheets: Edit target sessions")
    print("   • SESSION LOG Column K: Enter PIs worked on")
    print("   • Everything auto-calculates")
    print("   • Fully compatible with Microsoft 365 online")

if __name__ == '__main__':
    main()
