#!/usr/bin/env python3
"""
Script to update DG_MARE_TEAMS_SSF_Refinement_Tracker.xlsx with:
1. Labels column in JIRA DATA
2. New data from SSF_BACKLOG_EXPORT_10_11.xlsx
3. PI PROGRESS sheet
4. Enhanced EPIC PROGRESS with PI breakdown
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

def extract_pi_from_labels(labels_str):
    """Extract PI label from Labels string like 'Mobile, SSF_MOBILE_APP, SSF_PI#02'"""
    if not labels_str:
        return ""
    match = re.search(r'SSF_PI#\d+', str(labels_str))
    return match.group(0) if match else ""

def load_source_data(source_file):
    """Load data from SSF_BACKLOG_EXPORT_10_11.xlsx"""
    wb_source = openpyxl.load_workbook(source_file, data_only=True)
    ws_source = wb_source[wb_source.sheetnames[0]]
    
    # Headers are in row 4, data starts at row 5
    data = []
    for row in ws_source.iter_rows(min_row=5, values_only=True):
        if row[1]:  # if Issue key exists
            data.append({
                'Issue Type': row[0],
                'issueKey': row[1],
                'Issue id': row[2] if len(row) > 2 else None,
                'Summary': row[3],
                'Status': row[4],
                'Epic Link': row[5],
                'Story Points': row[6],
                'Labels': row[7] if len(row) > 7 else "",
                'Sprint': row[8] if len(row) > 8 else ""
            })
    
    wb_source.close()
    return data

def update_jira_data_sheet(ws, data):
    """Update JIRA DATA sheet with new structure and data"""
    # Update headers in row 2 to include Labels
    headers = ['Issue Type', 'Issue key', 'Issue id', 'Summary', 'Status', 
               'Epic Link', 'Story Points', 'Labels', 'Sprint', 'PI']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Clear existing data (row 3 onwards)
    ws.delete_rows(3, ws.max_row)
    
    # Insert new data
    for row_idx, item in enumerate(data, start=3):
        ws.cell(row=row_idx, column=1, value=item['Issue Type'])
        ws.cell(row=row_idx, column=2, value=item['issueKey'])
        ws.cell(row=row_idx, column=3, value=item['Issue id'])
        ws.cell(row=row_idx, column=4, value=item['Summary'])
        ws.cell(row=row_idx, column=5, value=item['Status'])
        ws.cell(row=row_idx, column=6, value=item['Epic Link'])
        ws.cell(row=row_idx, column=7, value=item['Story Points'])
        ws.cell(row=row_idx, column=8, value=item['Labels'])
        ws.cell(row=row_idx, column=9, value=item['Sprint'])
        
        # Column J: Extract PI from Labels
        pi_value = extract_pi_from_labels(item['Labels'])
        ws.cell(row=row_idx, column=10, value=pi_value)
    
    print(f"✅ Updated JIRA DATA sheet with {len(data)} issues")

def create_pi_progress_sheet(wb, ws_jira_data):
    """Create new PI PROGRESS sheet"""
    if 'PI PROGRESS' in wb.sheetnames:
        del wb['PI PROGRESS']
    
    ws = wb.create_sheet('PI PROGRESS', 2)  # Insert after DASHBOARD
    
    # Title row
    ws['A1'] = 'PI PROGRESS TRACKER'
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers row 2
    headers = ['PI Label', 'Total Stories', 'Refined (Ready)', 'To Refine (Open)', 
               'Other Status', '% Complete', 'Progress', 'Status']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get unique PIs from JIRA DATA
    pi_list = set()
    for row in ws_jira_data.iter_rows(min_row=3, values_only=True):
        pi_val = row[9] if len(row) > 9 else None  # Column J
        if pi_val and pi_val.startswith('SSF_PI#'):
            pi_list.add(pi_val)
    
    pi_list = sorted(list(pi_list))
    
    # Add PI rows with formulas
    for idx, pi in enumerate(pi_list, start=3):
        ws.cell(row=idx, column=1, value=pi)
        
        # B: Total Stories
        ws.cell(row=idx, column=2, value=f'=COUNTIF(\'JIRA DATA\'!J:J,A{idx})')
        
        # C: Refined (Ready)
        ws.cell(row=idx, column=3, value=f'=COUNTIFS(\'JIRA DATA\'!J:J,A{idx},\'JIRA DATA\'!E:E,"Ready")')
        
        # D: To Refine (Open)
        ws.cell(row=idx, column=4, value=f'=COUNTIFS(\'JIRA DATA\'!J:J,A{idx},\'JIRA DATA\'!E:E,"Open")')
        
        # E: Other Status
        ws.cell(row=idx, column=5, value=f'=B{idx}-C{idx}-D{idx}')
        
        # F: % Complete
        ws.cell(row=idx, column=6, value=f'=IF(B{idx}=0,0,C{idx}/B{idx})')
        ws.cell(row=idx, column=6).number_format = '0.0%'
        
        # G: Progress format (e.g., "3/35")
        ws.cell(row=idx, column=7, value=f'=C{idx}&"/"&B{idx}')
        
        # H: Status indicator
        ws.cell(row=idx, column=8, value=f'=IF(F{idx}>0.7,"🟢",IF(F{idx}>0.3,"🟡","🔴"))')
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 10
    
    print(f"✅ Created PI PROGRESS sheet with {len(pi_list)} PIs")

def update_epic_progress_sheet(ws):
    """Update EPIC PROGRESS sheet to include PI breakdown"""
    # Find the last column with data
    max_col = ws.max_column
    
    # Add new headers after existing columns
    new_headers = ['PI Filter', 'Refined in PI', 'Total in PI', 'PI Progress']
    start_col = max_col + 1
    
    for idx, header in enumerate(new_headers):
        cell = ws.cell(row=2, column=start_col + idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Add formulas for each epic row (starting from row 3)
    # Note: This assumes epics start at row 3
    # J: PI Filter (can be manually entered or set to specific PI)
    # K: Refined in PI
    # L: Total in PI
    # M: PI Progress format
    
    for row_idx in range(3, ws.max_row + 1):
        epic_cell = ws.cell(row=row_idx, column=1)  # Column A has Epic ID/Name
        if epic_cell.value:
            col_j = get_column_letter(start_col)
            col_k = get_column_letter(start_col + 1)
            col_l = get_column_letter(start_col + 2)
            col_m = get_column_letter(start_col + 3)
            
            # J: Can be filled manually with PI like "SSF_PI#02"
            # Leave blank for user to fill
            
            # K: Refined in PI (counts Ready status for Epic+PI)
            ws.cell(row=row_idx, column=start_col + 1, 
                   value=f'=IF({col_j}{row_idx}="","-",COUNTIFS(\'JIRA DATA\'!F:F,A{row_idx},\'JIRA DATA\'!J:J,{col_j}{row_idx},\'JIRA DATA\'!E:E,"Ready"))')
            
            # L: Total in PI
            ws.cell(row=row_idx, column=start_col + 2,
                   value=f'=IF({col_j}{row_idx}="","-",COUNTIFS(\'JIRA DATA\'!F:F,A{row_idx},\'JIRA DATA\'!J:J,{col_j}{row_idx}))')
            
            # M: Progress format
            ws.cell(row=row_idx, column=start_col + 3,
                   value=f'=IF({col_j}{row_idx}="","-",{col_k}{row_idx}&"/"&{col_l}{row_idx})')
    
    # Set column widths
    ws.column_dimensions[get_column_letter(start_col)].width = 12
    ws.column_dimensions[get_column_letter(start_col + 1)].width = 15
    ws.column_dimensions[get_column_letter(start_col + 2)].width = 12
    ws.column_dimensions[get_column_letter(start_col + 3)].width = 15
    
    print(f"✅ Updated EPIC PROGRESS sheet with PI tracking columns")

def update_dashboard_sheet(ws):
    """Update DASHBOARD to show PI-related metrics"""
    # Add a note about PI tracking
    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1, value="📊 PI PROGRESS")
    ws.cell(row=note_row, column=1).font = Font(size=14, bold=True, color="4472C4")
    
    ws.cell(row=note_row + 1, column=1, value="See PI PROGRESS sheet for detailed PI tracking →")
    ws.cell(row=note_row + 1, column=1).font = Font(italic=True)
    
    print(f"✅ Updated DASHBOARD with PI reference")

def main():
    source_file = 'SSF_BACKLOG_EXPORT_10_11.xlsx'
    tracker_file = 'DG_MARE_TEAMS_SSF_Refinement_Tracker.xlsx'
    output_file = 'DG_MARE_TEAMS_SSF_Refinement_Tracker.xlsx'
    
    print("🚀 Starting tracker update process...")
    
    # Load source data
    print(f"\n📂 Loading data from {source_file}...")
    data = load_source_data(source_file)
    print(f"   Found {len(data)} issues")
    
    # Load tracker workbook
    print(f"\n📂 Loading tracker from {tracker_file}...")
    wb = openpyxl.load_workbook(tracker_file)
    
    # Update JIRA DATA sheet
    print("\n📝 Updating JIRA DATA sheet...")
    ws_jira_data = wb['JIRA DATA']
    update_jira_data_sheet(ws_jira_data, data)
    
    # Create PI PROGRESS sheet
    print("\n📊 Creating PI PROGRESS sheet...")
    create_pi_progress_sheet(wb, ws_jira_data)
    
    # Update EPIC PROGRESS sheet
    print("\n📈 Updating EPIC PROGRESS sheet...")
    if 'EPIC PROGRESS' in wb.sheetnames:
        update_epic_progress_sheet(wb['EPIC PROGRESS'])
    
    # Update DASHBOARD
    print("\n📋 Updating DASHBOARD...")
    if 'DASHBOARD' in wb.sheetnames:
        update_dashboard_sheet(wb['DASHBOARD'])
    
    # Save the updated workbook
    print(f"\n💾 Saving to {output_file}...")
    wb.save(output_file)
    wb.close()
    
    print("\n" + "="*60)
    print("✅ TRACKER UPDATE COMPLETE!")
    print("="*60)
    print(f"\n📊 Summary:")
    print(f"   • Imported {len(data)} issues from JIRA export")
    print(f"   • Added Labels column to JIRA DATA")
    print(f"   • Created PI PROGRESS sheet")
    print(f"   • Enhanced EPIC PROGRESS with PI breakdown")
    print(f"   • Updated DASHBOARD with PI reference")
    print(f"\n📁 Output file: {output_file}")
    print("\n🎯 Next steps:")
    print("   1. Open the file in Excel or Google Sheets")
    print("   2. Check PI PROGRESS sheet for PI-level metrics")
    print("   3. In EPIC PROGRESS, fill column J with PI labels (e.g., 'SSF_PI#02')")
    print("   4. Verify formulas calculate correctly")
    print("   5. Update SESSION LOG with refinement session data")

if __name__ == '__main__':
    main()
