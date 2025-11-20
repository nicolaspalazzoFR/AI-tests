#!/usr/bin/env python3
"""
Enhanced Tracker Update Script:
1. Fix EPIC PROGRESS (use Epic Names instead of IDs)
2. Create PI-level session logs (filtered from master)
3. Create PI-level burndown charts with flexible session counts
4. Create optional dynamic Epic burndown control
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference
import re

def get_unique_epics(ws_jira_data):
    """Get list of unique epic names from JIRA DATA"""
    epics = set()
    for row in ws_jira_data.iter_rows(min_row=3, values_only=True):
        epic = row[5]  # Column F - Epic Link
        if epic and epic != "Epic Link":
            epics.add(epic)
    return sorted(list(epics))

def get_unique_pis(ws_jira_data):
    """Get list of unique PIs from JIRA DATA"""
    pis = set()
    for row in ws_jira_data.iter_rows(min_row=3, values_only=True):
        pi = row[9]  # Column J - PI
        if pi and pi.startswith('SSF_PI#'):
            pis.add(pi)
    return sorted(list(pis))

def fix_epic_progress(ws_epic, epics):
    """Fix EPIC PROGRESS to use Epic Names instead of IDs"""
    print("\n🔧 Fixing EPIC PROGRESS Sheet...")
    
    # Clear existing epic rows (keep headers)
    if ws_epic.max_row > 2:
        ws_epic.delete_rows(3, ws_epic.max_row - 2)
    
    # Add each epic
    for idx, epic_name in enumerate(epics, start=3):
        ws_epic.cell(row=idx, column=1, value=epic_name)  # Column A: Epic Name
        ws_epic.cell(row=idx, column=2, value=epic_name)  # Column B: Epic Name (duplicate for display)
        
        # Column C: Total
        ws_epic.cell(row=idx, column=3, value=f"=COUNTIF('JIRA DATA'!F:F,A{idx})")
        
        # Column D: Open
        ws_epic.cell(row=idx, column=4, value=f"=COUNTIFS('JIRA DATA'!F:F,A{idx},'JIRA DATA'!E:E,\"Open\")")
        
        # Column E: Ready
        ws_epic.cell(row=idx, column=5, value=f"=COUNTIFS('JIRA DATA'!F:F,A{idx},'JIRA DATA'!E:E,\"Ready\")")
        
        # Column F: Other
        ws_epic.cell(row=idx, column=6, value=f"=C{idx}-D{idx}-E{idx}")
        
        # Column G: % Complete
        ws_epic.cell(row=idx, column=7, value=f"=IF(C{idx}=0,0,E{idx}/C{idx})")
        ws_epic.cell(row=idx, column=7).number_format = '0.0%'
        
        # Column H: Priority (leave blank for manual entry)
        # Column I: Sprint (leave blank for manual entry)
        
        # Column J: PI Filter (leave blank for manual entry)
        
        # Column K: Refined in PI
        ws_epic.cell(row=idx, column=11, value=f"=IF(J{idx}=\"\",\"-\",COUNTIFS('JIRA DATA'!F:F,A{idx},'JIRA DATA'!J:J,J{idx},'JIRA DATA'!E:E,\"Ready\"))")
        
        # Column L: Total in PI
        ws_epic.cell(row=idx, column=12, value=f"=IF(J{idx}=\"\",\"-\",COUNTIFS('JIRA DATA'!F:F,A{idx},'JIRA DATA'!J:J,J{idx}))")
        
        # Column M: PI Progress
        ws_epic.cell(row=idx, column=13, value=f"=IF(J{idx}=\"\",\"-\",K{idx}&\"/\"&L{idx})")
    
    print(f"   ✅ Fixed {len(epics)} epics with proper formulas")

def create_pi_session_log(wb, pi_label, ws_master_log, ws_jira_data):
    """Create filtered session log sheet for a specific PI"""
    sheet_name = f"{pi_label} SESSION LOG"
    
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name)
    
    # Title
    ws['A1'] = f'{pi_label} REFINEMENT SESSION LOG'
    ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('A1:I1')
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    headers = ['Session #', 'Sprint', 'Date', 'PI(s) Worked', 'Target Stories', 
               'Stories Refined', 'Cumulative', 'Remaining', 'Velocity']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Get total stories for this PI
    pi_total_formula = f'=COUNTIF(\'JIRA DATA\'!J:J,"{pi_label}")'
    
    # Add data rows with formulas that filter from master
    for row_idx in range(3, 20):  # Up to 17 sessions
        session_num = row_idx - 2
        
        # A: Session #
        ws.cell(row=row_idx, column=1, value=session_num)
        
        # B: Sprint (from master)
        ws.cell(row=row_idx, column=2, value=f"='SESSION LOG'!B{row_idx}")
        
        # C: Date (from master)
        ws.cell(row=row_idx, column=3, value=f"='SESSION LOG'!C{row_idx}")
        
        # D: PI(s) Worked (from master - new column we'll add)
        ws.cell(row=row_idx, column=4, value=f"='SESSION LOG'!K{row_idx}")
        
        # E: Target Stories
        ws.cell(row=row_idx, column=5, value=f"='SESSION LOG'!E{row_idx}")
        
        # F: Stories Refined (only count if this PI was worked on)
        # Check if column K in master contains this PI label
        ws.cell(row=row_idx, column=6, 
               value=f'=IF(ISNUMBER(SEARCH("{pi_label}",\'SESSION LOG\'!K{row_idx})),\'SESSION LOG\'!F{row_idx},0)')
        
        # G: Cumulative
        ws.cell(row=row_idx, column=7, value=f"=SUM($F$3:F{row_idx})")
        
        # H: Remaining
        ws.cell(row=row_idx, column=8, value=f"={pi_total_formula}-G{row_idx}")
        
        # I: Velocity
        if row_idx == 3:
            ws.cell(row=row_idx, column=9, value=f"=F{row_idx}")
        else:
            ws.cell(row=row_idx, column=9, value=f"=AVERAGE($F$3:F{row_idx})")
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 12
    ws.column_dimensions['H'].width = 12
    ws.column_dimensions['I'].width = 10
    
    return ws

def create_pi_burndown(wb, pi_label, ws_jira_data):
    """Create burndown chart sheet for a specific PI with flexible session count"""
    sheet_name = f"{pi_label} BURNDOWN"
    
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name)
    
    # Configuration cells
    ws['A1'] = 'TARGET SESSIONS:'
    ws['A1'].font = Font(bold=True)
    ws['B1'] = 15  # Default target sessions - USER CAN EDIT
    ws['B1'].font = Font(bold=True, color="FF0000")
    ws['B1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    ws['A2'] = 'TOTAL STORIES:'
    ws['A2'].font = Font(bold=True)
    ws['B2'] = f'=COUNTIF(\'JIRA DATA\'!J:J,"{pi_label}")'
    ws['B2'].font = Font(bold=True)
    
    # Title
    ws['D1'] = f'{pi_label} BURNDOWN CHART'
    ws['D1'].font = Font(size=14, bold=True, color="FFFFFF")
    ws['D1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    ws.merge_cells('D1:H1')
    ws['D1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Headers
    ws['A4'] = 'Session #'
    ws['B4'] = 'Sprint'
    ws['C4'] = 'IDEAL Remaining'
    ws['D4'] = 'ACTUAL Remaining'
    ws['E4'] = 'Variance'
    
    for cell in [ws['A4'], ws['B4'], ws['C4'], ws['D4'], ws['E4']]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Data rows
    session_log_sheet = f"{pi_label} SESSION LOG"
    
    for row_idx in range(5, 22):  # Up to 17 sessions
        session_num = row_idx - 4
        
        # A: Session #
        ws.cell(row=row_idx, column=1, value=session_num - 1)  # Start at 0
        
        # B: Sprint
        ws.cell(row=row_idx, column=2, value=f"='{session_log_sheet}'!B{row_idx-2}")
        
        # C: IDEAL Remaining (flexible based on target sessions)
        ws.cell(row=row_idx, column=3, 
               value=f"=$B$2-($B$2/$B$1)*A{row_idx}")
        
        # D: ACTUAL Remaining
        ws.cell(row=row_idx, column=4, value=f"='{session_log_sheet}'!H{row_idx-2}")
        
        # E: Variance
        ws.cell(row=row_idx, column=5, value=f"=D{row_idx}-C{row_idx}")
    
    # Create chart
    chart = LineChart()
    chart.title = f"{pi_label} - Ideal vs Actual Refinement Progress"
    chart.style = 10
    chart.y_axis.title = "Stories Remaining"
    chart.x_axis.title = "Session #"
    chart.height = 10
    chart.width = 20
    
    # Data for chart
    ideal_data = Reference(ws, min_col=3, min_row=4, max_row=21)
    actual_data = Reference(ws, min_col=4, min_row=4, max_row=21)
    sessions = Reference(ws, min_col=1, min_row=5, max_row=21)
    
    chart.add_data(ideal_data, titles_from_data=True)
    chart.add_data(actual_data, titles_from_data=True)
    chart.set_categories(sessions)
    
    # Style the series
    chart.series[0].graphicalProperties.line.solidFill = "4472C4"  # Blue for ideal
    chart.series[0].graphicalProperties.line.width = 2.5
    chart.series[1].graphicalProperties.line.solidFill = "FF0000"  # Red for actual
    chart.series[1].graphicalProperties.line.width = 2.5
    
    ws.add_chart(chart, "G4")
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    
    print(f"   ✅ Created {sheet_name} with flexible session count")

def add_pi_column_to_session_log(ws_session):
    """Add PI(s) Worked column to master SESSION LOG"""
    print("\n🔧 Enhancing SESSION LOG...")
    
    # Add header in column K
    ws_session.cell(row=2, column=11, value="PI(s) Worked")
    ws_session.cell(row=2, column=11).font = Font(bold=True, color="FFFFFF")
    ws_session.cell(row=2, column=11).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws_session.cell(row=2, column=11).alignment = Alignment(horizontal='center', vertical='center')
    
    # Add note in first data row
    ws_session.cell(row=3, column=11, value="Enter PI labels: SSF_PI#01, SSF_PI#02, etc.")
    ws_session.cell(row=3, column=11).font = Font(italic=True, color="999999")
    
    ws_session.column_dimensions['K'].width = 25
    
    print("   ✅ Added 'PI(s) Worked' column to SESSION LOG")

def main():
    tracker_file = 'DG_MARE_TEAMS_SSF_Refinement_Tracker.xlsx'
    
    print("🚀 Starting Enhanced Tracker Fix...")
    print("="*60)
    
    # Load workbook
    print(f"\n📂 Loading {tracker_file}...")
    wb = openpyxl.load_workbook(tracker_file)
    
    ws_jira_data = wb['JIRA DATA']
    ws_epic = wb['EPIC PROGRESS']
    ws_session = wb['SESSION LOG']
    
    # Get unique epics and PIs
    print("\n📊 Analyzing data...")
    epics = get_unique_epics(ws_jira_data)
    pis = get_unique_pis(ws_jira_data)
    
    print(f"   Found {len(epics)} unique epics: {epics[:5]}{'...' if len(epics) > 5 else ''}")
    print(f"   Found {len(pis)} unique PIs: {pis}")
    
    # 1. Fix EPIC PROGRESS
    fix_epic_progress(ws_epic, epics)
    
    # 2. Add PI column to master SESSION LOG
    add_pi_column_to_session_log(ws_session)
    
    # 3. Create PI-level session logs and burndowns
    print(f"\n📈 Creating PI-level tracking...")
    for pi in pis:
        print(f"\n   Processing {pi}...")
        create_pi_session_log(wb, pi, ws_session, ws_jira_data)
        create_pi_burndown(wb, pi, ws_jira_data)
    
    # 4. Update existing BURNDOWN CHART DATA to have flexible sessions
    print(f"\n🔧 Updating overall BURNDOWN CHART DATA...")
    ws_burndown = wb['BURNDOWN CHART DATA']
    
    # Add target sessions cell
    ws_burndown.insert_rows(1)
    ws_burndown['A1'] = 'TARGET SESSIONS:'
    ws_burndown['A1'].font = Font(bold=True)
    ws_burndown['B1'] = 15
    ws_burndown['B1'].font = Font(bold=True, color="FF0000")
    ws_burndown['B1'].fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Update ideal line formula to use flexible sessions
    for row_idx in range(5, ws_burndown.max_row + 1):
        cell = ws_burndown.cell(row=row_idx, column=3)
        if cell.value and '15' in str(cell.value):
            # Replace hardcoded 15 with reference to B1
            cell.value = cell.value.replace('/15', '/$B$1')
    
    print("   ✅ Updated with flexible session count")
    
    # Save
    print(f"\n💾 Saving changes...")
    wb.save(tracker_file)
    wb.close()
    
    print("\n" + "="*60)
    print("✅ ENHANCED TRACKER UPDATE COMPLETE!")
    print("="*60)
    print(f"\n📊 Summary:")
    print(f"   • Fixed EPIC PROGRESS with {len(epics)} epics")
    print(f"   • Added PI(s) Worked column to SESSION LOG")
    print(f"   • Created {len(pis)} PI SESSION LOG sheets")
    print(f"   • Created {len(pis)} PI BURNDOWN sheets")
    print(f"   • All burndowns have FLEXIBLE session counts")
    
    print(f"\n🎯 Next Steps:")
    print(f"   1. Open the tracker file")
    print(f"   2. In SESSION LOG column K, enter PI labels for each session")
    print(f"      Example: 'SSF_PI#02' or 'SSF_PI#01, SSF_PI#02'")
    print(f"   3. In each PI BURNDOWN sheet, edit cell B1 to change target sessions")
    print(f"   4. Enter Stories Refined in SESSION LOG column F after each session")
    print(f"   5. All charts will auto-update!")
    
    print(f"\n📝 Key Features:")
    print(f"   • EPIC PROGRESS now shows correct counts")
    print(f"   • PI-level burndowns filter from master session log")
    print(f"   • Change target sessions in any burndown (cell B1)")
    print(f"   • Ideal line recalculates automatically")

if __name__ == '__main__':
    main()
