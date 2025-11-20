#!/usr/bin/env python3
"""
Create SSF Refinement Tracker & Session Planner Excel file
Compatible with Google Sheets import - All formulas corrected
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, Reference

def style_header(cell, bg='366092', fg='FFFFFF'):
    cell.font = Font(bold=True, color=fg, size=11)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def main():
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Create DASHBOARD
    ws = wb.create_sheet('DASHBOARD')
    ws['A1'] = 'REFINEMENT TRACKER - DASHBOARD'
    style_header(ws['A1'])
    ws.merge_cells('A1:G1')
    
    ws['A2'], ws['B2'], ws['C2'], ws['D2'] = 'Last Updated', '=TODAY()', 'Current Sprint', 1
    ws['A4'] = 'QUICK METRICS'
    ws['A6'], ws['C6'], ws['E6'] = 'Total Stories', 91, 'Status'
    ws['F6'] = '=IF(F9="","-",IF(F9>=6.5,"🟢 On Track",IF(F9>=4,"🟡 At Risk","🔴 Critical")))'
    ws['A7'], ws['C7'], ws['E7'] = 'To Refine (Open)', 84, 'Risk Level'
    ws['F7'] = '=IF(C10>60,"🔴 HIGH",IF(C10>40,"🟡 MEDIUM","🟢 LOW"))'
    ws['A8'], ws['C8'] = 'Already Refined (Ready)', 7
    ws['A9'], ws['C9'] = 'Completed So Far', '=SUM(\'SESSION LOG\'!F:F)'
    ws['A10'], ws['C10'] = 'Remaining', '=IF(ISBLANK(C9),84,84-C9)'
    ws['A11'], ws['C11'] = 'Progress %', '=IF(C9=0,0,C9/84*100)'
    ws['E9'], ws['F9'] = 'Current Velocity', '=IF(COUNTA(\'SESSION LOG\'!F2:F16)=0,0,AVERAGE(\'SESSION LOG\'!F2:F16))'
    ws['E10'], ws['F10'] = 'Target Velocity', 6.5
    ws['E11'], ws['F11'] = 'Projected Completion', '=IF(OR(F9=0,F9=""),"TBD","Sprint "&ROUNDUP(C10/F9/3+1,0))'
    
    ws['A13'] = 'SESSION SUMMARY (LAST 3)'
    ws.merge_cells('A13:G13')
    for col, hdr in enumerate(['Session', 'Sprint', 'Date', 'Stories Refined', 'Cumulative', 'Status'], start=2):
        ws.cell(14, col, hdr)
        style_header(ws.cell(14, col))
    
    ws['B15'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=1,"",INDEX(\'SESSION LOG\'!A:A,COUNTA(\'SESSION LOG\'!A:A)))'
    ws['C15'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=1,"",INDEX(\'SESSION LOG\'!B:B,COUNTA(\'SESSION LOG\'!A:A)))'
    ws['D15'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=1,"",INDEX(\'SESSION LOG\'!C:C,COUNTA(\'SESSION LOG\'!A:A)))'
    ws['E15'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=1,"",INDEX(\'SESSION LOG\'!F:F,COUNTA(\'SESSION LOG\'!A:A)))'
    ws['F15'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=1,"",INDEX(\'SESSION LOG\'!G:G,COUNTA(\'SESSION LOG\'!A:A)))'
    ws['G15'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=1,"",INDEX(\'SESSION LOG\'!J:J,COUNTA(\'SESSION LOG\'!A:A)))'
    
    ws['B16'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=2,"",INDEX(\'SESSION LOG\'!A:A,COUNTA(\'SESSION LOG\'!A:A)-1))'
    ws['C16'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=2,"",INDEX(\'SESSION LOG\'!B:B,COUNTA(\'SESSION LOG\'!A:A)-1))'
    ws['D16'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=2,"",INDEX(\'SESSION LOG\'!C:C,COUNTA(\'SESSION LOG\'!A:A)-1))'
    ws['E16'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=2,"",INDEX(\'SESSION LOG\'!F:F,COUNTA(\'SESSION LOG\'!A:A)-1))'
    ws['F16'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=2,"",INDEX(\'SESSION LOG\'!G:G,COUNTA(\'SESSION LOG\'!A:A)-1))'
    ws['G16'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=2,"",INDEX(\'SESSION LOG\'!J:J,COUNTA(\'SESSION LOG\'!A:A)-1))'
    
    ws['B17'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=3,"",INDEX(\'SESSION LOG\'!A:A,COUNTA(\'SESSION LOG\'!A:A)-2))'
    ws['C17'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=3,"",INDEX(\'SESSION LOG\'!B:B,COUNTA(\'SESSION LOG\'!A:A)-2))'
    ws['D17'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=3,"",INDEX(\'SESSION LOG\'!C:C,COUNTA(\'SESSION LOG\'!A:A)-2))'
    ws['E17'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=3,"",INDEX(\'SESSION LOG\'!F:F,COUNTA(\'SESSION LOG\'!A:A)-2))'
    ws['F17'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=3,"",INDEX(\'SESSION LOG\'!G:G,COUNTA(\'SESSION LOG\'!A:A)-2))'
    ws['G17'] = '=IF(COUNTA(\'SESSION LOG\'!A:A)<=3,"",INDEX(\'SESSION LOG\'!J:J,COUNTA(\'SESSION LOG\'!A:A)-2))'
    
    ws['A19'] = 'WEEKLY REPORT (Auto-Generated)'
    ws.merge_cells('A19:G19')
    ws['A20'] = 'Report Text'
    ws.merge_cells('B20:G25')
    ws['B20'] = '="REFINEMENT PROGRESS - Week of "&TEXT(TODAY(),"MMM DD, YYYY")&CHAR(10)&CHAR(10)&"✅ STORIES REFINED THIS WEEK"&CHAR(10)&"   • Sessions conducted: "&COUNTIFS(\'SESSION LOG\'!C:C,">="&TODAY()-7,\'SESSION LOG\'!C:C,"<="&TODAY())&CHAR(10)&"   • Stories moved to Ready: "&SUMIFS(\'SESSION LOG\'!F:F,\'SESSION LOG\'!C:C,">="&TODAY()-7,\'SESSION LOG\'!C:C,"<="&TODAY())&CHAR(10)&"   • Cumulative progress: "&C9&"/84 ("&ROUND(C11,0)&"%)"&CHAR(10)&CHAR(10)&"📈 VELOCITY METRICS"&CHAR(10)&"   • Current velocity: "&ROUND(F9,1)&" stories/session"&CHAR(10)&"   • Target velocity: 6.5 stories/session"&CHAR(10)&"   • Status: "&F6&CHAR(10)&CHAR(10)&"📅 TIMELINE"&CHAR(10)&"   • Sessions remaining: "&(15-COUNTA(\'SESSION LOG\'!F2:F16))&CHAR(10)&"   • Projected completion: "&F11&CHAR(10)&"   • On track for Sprint 5: "&IF(F11="Sprint 5","YES ✅",IF(F11="Sprint 6","ALMOST ⚠️","NO ❌"))'
    ws['B20'].alignment = Alignment(wrap_text=True, vertical='top')
    
    for col, w in zip(['A','B','C','D','E','F','G'], [25,12,12,12,20,20,12]):
        ws.column_dimensions[col].width = w
    
    # Create SESSION LOG
    ws = wb.create_sheet('SESSION LOG')
    ws['A1'] = 'SESSION LOG - Data Entry (Update After Each Session)'
    style_header(ws['A1'])
    ws.merge_cells('A1:K1')
    
    for col, hdr in enumerate(['Session #', 'Sprint', 'Date', 'Day', 'Target Stories', 'Stories Refined', 
                               'Cumulative', 'Remaining', 'Velocity', 'Status', 'Notes/Blockers'], start=1):
        ws.cell(2, col, hdr)
        style_header(ws.cell(2, col))
    
    sessions = [(1,1,'2025-11-05','Tuesday',7,0,'Learning session'),(2,1,'2025-11-07','Thursday',7,1,'Team alignment'),
                (3,1,'2025-11-12','Tuesday',7,'',''),(4,2,'2025-11-19','Tuesday',7,'',''),(5,2,'2025-11-21','Thursday',7,'',''),
                (6,2,'2025-11-26','Tuesday',7,'',''),(7,3,'2025-12-03','Tuesday',7,'',''),(8,3,'2025-12-05','Thursday',7,'',''),
                (9,3,'2025-12-10','Tuesday',7,'',''),(10,4,'2025-12-17','Tuesday',7,'',''),(11,4,'2025-12-19','Thursday',7,'',''),
                (12,4,'2025-12-24','Tuesday',7,'',''),(13,5,'2026-01-07','Tuesday',7,'',''),(14,5,'2026-01-09','Thursday',7,'',''),
                (15,5,'2026-01-14','Tuesday',7,'','')]
    
    for idx, (s,sp,d,dy,t,r,n) in enumerate(sessions, start=3):
        ws[f'A{idx}'], ws[f'B{idx}'], ws[f'C{idx}'], ws[f'D{idx}'], ws[f'E{idx}'] = s, sp, d, dy, t
        if r != '': ws[f'F{idx}'] = r
        ws[f'G{idx}'] = f'=SUM($F$3:F{idx})'
        ws[f'H{idx}'] = f'=84-G{idx}'
        ws[f'I{idx}'] = f'=F{idx}'
        ws[f'J{idx}'] = f'=IF(ISBLANK(I{idx}),"",IF(I{idx}>=6.5,"🟢",IF(I{idx}>=4,"🟡","🔴")))'
        ws[f'K{idx}'] = n
    
    ws['A19'] = 'INSTRUCTIONS: Update ONLY column F after each session. Everything else auto-calculates!'
    ws['A19'].font = Font(bold=True, color='FF0000')
    
    for col, w in zip(['A','B','C','D','E','F','G','H','I','J','K'], [10,8,12,10,14,14,12,12,10,8,30]):
        ws.column_dimensions[col].width = w
    
    # Create BURNDOWN CHART DATA
    ws = wb.create_sheet('BURNDOWN CHART DATA')
    ws['A1'] = 'BURNDOWN CHART DATA'
    style_header(ws['A1'])
    ws.merge_cells('A1:F1')
    
    for col, hdr in enumerate(['Session #', 'Sprint', 'Stories Remaining (Ideal)', 'Stories Remaining (Actual)', 'Variance', 'Status'], start=1):
        ws.cell(2, col, hdr)
        style_header(ws.cell(2, col))
    
    ws['A3'], ws['B3'], ws['C3'], ws['D3'], ws['E3'], ws['F3'] = 0, 1, 84, 84, 0, 'Start'
    
    for row in range(4, 19):
        sr = row - 1
        ws[f'A{row}'] = f'=INDEX(\'SESSION LOG\'!A:A,{sr})'
        ws[f'B{row}'] = f'=INDEX(\'SESSION LOG\'!B:B,{sr})'
        ws[f'C{row}'] = f'=84-5.6*({row-3})'
        ws[f'D{row}'] = f'=INDEX(\'SESSION LOG\'!H:H,{sr})'
        ws[f'E{row}'] = f'=D{row}-C{row}'
        ws[f'F{row}'] = f'=IF(E{row}>=0,"🟢",IF(E{row}>=-5,"🟡","🔴"))'
    
    for col in ['A','B','C','D','E','F']:
        ws.column_dimensions[col].width = 18
    
    # Create BURNUP CHART DATA
    ws = wb.create_sheet('BURNUP CHART DATA')
    ws['A1'] = 'BURNUP CHART DATA'
    style_header(ws['A1'])
    ws.merge_cells('A1:F1')
    
    for col, hdr in enumerate(['Session #', 'Sprint', 'Stories Refined (Target)', 'Stories Refined (Actual)', 'Total Scope', 'Gap'], start=1):
        ws.cell(2, col, hdr)
        style_header(ws.cell(2, col))
    
    ws['A3'], ws['B3'], ws['C3'], ws['D3'], ws['E3'], ws['F3'] = 0, 1, 0, 0, 84, 0
    
    for row in range(4, 19):
        sr = row - 1
        ws[f'A{row}'] = f'=INDEX(\'SESSION LOG\'!A:A,{sr})'
        ws[f'B{row}'] = f'=INDEX(\'SESSION LOG\'!B:B,{sr})'
        ws[f'C{row}'] = f'=5.6*({row-3})'
        ws[f'D{row}'] = f'=INDEX(\'SESSION LOG\'!G:G,{sr})'
        ws[f'E{row}'] = 84
        ws[f'F{row}'] = f'=D{row}-C{row}'
    
    for col in ['A','B','C','D','E','F']:
        ws.column_dimensions[col].width = 18
    
    # Add BURNDOWN CHART to the sheet
    burndown_ws = wb['BURNDOWN CHART DATA']
    chart_burndown = LineChart()
    chart_burndown.title = "Burndown Chart - Stories Remaining"
    chart_burndown.style = 10
    chart_burndown.y_axis.title = 'Stories Remaining'
    chart_burndown.x_axis.title = 'Session #'
    chart_burndown.height = 10
    chart_burndown.width = 20
    
    # Add data for burndown - Ideal line
    data_ideal = Reference(burndown_ws, min_col=3, min_row=2, max_row=18)
    chart_burndown.add_data(data_ideal, titles_from_data=True)
    
    # Add data for burndown - Actual line
    data_actual = Reference(burndown_ws, min_col=4, min_row=2, max_row=18)
    chart_burndown.add_data(data_actual, titles_from_data=True)
    
    # Set categories (Session numbers)
    cats = Reference(burndown_ws, min_col=1, min_row=3, max_row=18)
    chart_burndown.set_categories(cats)
    
    # Style the series
    s1 = chart_burndown.series[0]
    s1.graphicalProperties.line.solidFill = "4472C4"  # Blue for Ideal
    s1.graphicalProperties.line.width = 30000
    s1.smooth = True
    
    s2 = chart_burndown.series[1]
    s2.graphicalProperties.line.solidFill = "ED7D31"  # Orange for Actual
    s2.graphicalProperties.line.width = 30000
    s2.smooth = True
    
    # Add chart to sheet
    burndown_ws.add_chart(chart_burndown, "H3")
    
    # Add BURNUP CHART to the sheet
    burnup_ws = wb['BURNUP CHART DATA']
    chart_burnup = LineChart()
    chart_burnup.title = "Burnup Chart - Stories Refined"
    chart_burnup.style = 10
    chart_burnup.y_axis.title = 'Stories Refined'
    chart_burnup.x_axis.title = 'Session #'
    chart_burnup.height = 10
    chart_burnup.width = 20
    
    # Add data for burnup - Target line
    data_target = Reference(burnup_ws, min_col=3, min_row=2, max_row=18)
    chart_burnup.add_data(data_target, titles_from_data=True)
    
    # Add data for burnup - Actual line
    data_actual_burnup = Reference(burnup_ws, min_col=4, min_row=2, max_row=18)
    chart_burnup.add_data(data_actual_burnup, titles_from_data=True)
    
    # Add data for burnup - Total Scope line
    data_scope = Reference(burnup_ws, min_col=5, min_row=2, max_row=18)
    chart_burnup.add_data(data_scope, titles_from_data=True)
    
    # Set categories (Session numbers)
    cats_burnup = Reference(burnup_ws, min_col=1, min_row=3, max_row=18)
    chart_burnup.set_categories(cats_burnup)
    
    # Style the series
    s1_burnup = chart_burnup.series[0]
    s1_burnup.graphicalProperties.line.solidFill = "4472C4"  # Blue for Target
    s1_burnup.graphicalProperties.line.width = 30000
    s1_burnup.smooth = True
    
    s2_burnup = chart_burnup.series[1]
    s2_burnup.graphicalProperties.line.solidFill = "70AD47"  # Green for Actual
    s2_burnup.graphicalProperties.line.width = 30000
    s2_burnup.smooth = True
    
    s3_burnup = chart_burnup.series[2]
    s3_burnup.graphicalProperties.line.solidFill = "A5A5A5"  # Gray for Total Scope
    s3_burnup.graphicalProperties.line.width = 20000
    s3_burnup.graphicalProperties.line.dashStyle = "dash"
    
    # Add chart to sheet
    burnup_ws.add_chart(chart_burnup, "H3")
    
    # Create EPIC PROGRESS
    ws = wb.create_sheet('EPIC PROGRESS')
    ws['A1'] = 'EPIC PROGRESS'
    style_header(ws['A1'])
    ws.merge_cells('A1:I1')
    
    for col, hdr in enumerate(['Epic ID', 'Epic Name', 'Total', 'Open', 'Ready', 'Other', '% Complete', 'Priority', 'Sprint'], start=1):
        ws.cell(2, col, hdr)
        style_header(ws.cell(2, col))
    
    epics = [('SSF-123','Fishing Operations',14,13,1,0,'HIGH','1-2'),('SSF-96','Login & Session',12,6,6,0,'CRITICAL','1'),
             ('SSF-122','Departure & Catches',10,10,0,0,'HIGH','2-3'),('SSF-93','Gear Management',10,9,1,0,'MEDIUM','2-3'),
             ('SSF-95','Map & Location',8,6,2,0,'MEDIUM','3-4'),('SSF-84','Technical Setup',8,3,5,0,'HIGH','1'),
             ('SSF-165','Backend Infrastructure',3,3,0,0,'MEDIUM','4-5'),('SSF-166','Internationalization',4,3,1,0,'LOW','5'),
             ('SSF-162','Menu & Navigation',3,2,1,0,'MEDIUM','3'),('SSF-94','Home & Widgets',5,4,1,0,'MEDIUM','2'),
             ('SSF-86','Tracking',6,1,5,0,'HIGH','1'),('SSF-97','CI/CD & Deployment',5,5,0,0,'MEDIUM','4-5'),
             ('SSF-132','FMC Integration',1,1,0,0,'MEDIUM','4')]
    
    for idx, (eid,name,tot,opn,rdy,oth,pri,spr) in enumerate(epics, start=3):
        ws[f'A{idx}'], ws[f'B{idx}'], ws[f'C{idx}'], ws[f'D{idx}'] = eid, name, tot, opn
        ws[f'E{idx}'], ws[f'F{idx}'], ws[f'H{idx}'], ws[f'I{idx}'] = rdy, oth, pri, spr
        ws[f'G{idx}'] = f'=IF(C{idx}=0,0,E{idx}/C{idx}*100)'
    
    for col, w in zip(['A','B','C','D','E','F','G','H','I'], [12,25,12,8,8,8,12,12,14]):
        ws.column_dimensions[col].width = w
    
    # Create SESSION CALENDAR
    ws = wb.create_sheet('SESSION CALENDAR')
    ws['A1'] = 'SESSION CALENDAR - All 13 Remaining Sessions'
    style_header(ws['A1'])
    ws.merge_cells('A1:K1')
    
    for col, hdr in enumerate(['Session #', 'Sprint', 'Week', 'Date', 'Day', 'Time', 'Duration', 
                               'Target', 'Prep Deadline', 'Status', 'Notes'], start=1):
        ws.cell(2, col, hdr)
        style_header(ws.cell(2, col))
    
    cal = [(3,1,2,'2025-11-12','Tue','14:00',60,7,'2025-11-10','🟡 This Week','Quick wins'),
           (4,2,1,'2025-11-19','Tue','14:00',60,7,'2025-11-17','🟢 Upcoming','Gear Backend'),
           (5,2,1,'2025-11-21','Thu','14:00',60,7,'2025-11-19','🟢 Upcoming','Gear Frontend'),
           (6,2,2,'2025-11-26','Tue','14:00',60,7,'2025-11-24','🟢 Upcoming','Map & Location'),
           (7,3,1,'2025-12-03','Tue','14:00',60,7,'2025-12-01','🟢 Upcoming','Departure Ops'),
           (8,3,1,'2025-12-05','Thu','14:00',60,7,'2025-12-03','🟢 Upcoming','Catches'),
           (9,3,2,'2025-12-10','Tue','14:00',60,7,'2025-12-08','🟢 Upcoming','Backend'),
           (10,4,1,'2025-12-17','Tue','14:00',60,7,'2025-12-15','🟢 Upcoming','Authentication'),
           (11,4,1,'2025-12-19','Thu','14:00',60,7,'2025-12-17','🟢 Upcoming','Infrastructure'),
           (12,4,2,'2025-12-24','Tue','14:00',60,7,'2025-12-22','🟢 Upcoming','Buffer'),
           (13,5,1,'2026-01-07','Tue','14:00',60,7,'2026-01-05','🟢 Upcoming','Final push'),
           (14,5,1,'2026-01-09','Thu','14:00',60,7,'2026-01-07','🟢 Upcoming','Final push'),
           (15,5,2,'2026-01-14','Tue','14:00',60,7,'2026-01-12','🟢 Upcoming','Completion')]
    
    for idx, data in enumerate(cal, start=3):
        for col, val in enumerate(data, start=1):
            ws.cell(idx, col, val)
    
    for idx, w in enumerate([10,8,6,12,10,8,12,12,16,14,20], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w
    
    # Save
    wb.save('SSF_Refinement_Tracker_Session_Planner.xlsx')
    print("✅ Excel file created: SSF_Refinement_Tracker_Session_Planner.xlsx")
    print("📊 6 sheets: DASHBOARD, SESSION LOG, BURNDOWN, BURNUP, EPIC PROGRESS, SESSION CALENDAR")
    print("🔧 All formulas corrected for Google Sheets compatibility")
    print("\n📋 To use: Import to Google Sheets and update SESSION LOG column F after each session")

if __name__ == '__main__':
    main()
